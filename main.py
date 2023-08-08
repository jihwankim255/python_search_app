import xlrd
import openpyxl
from excelSearch import Ui_MainWindow
from PyQt5.QtWidgets import QMainWindow, QApplication, QTableWidgetItem,QLabel,QPushButton, QFileDialog, QMessageBox, QErrorMessage
from PyQt5.QtCore import Qt, QSettings
import openpyxl
import sys
import time
from Utils import sheet_search, buttonStart_pressed2, result_file_list
from openpyxl.utils import get_column_letter
import pyexcel as pe
import webbrowser


class MyMainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MyMainWindow, self).__init__(parent)
        self.setupUi(self)
        self.fileAButton.clicked.connect(self.buttonA_pressed)
        self.fileBButton.clicked.connect(self.buttonB_pressed)
        self.startButton.clicked.connect(self.buttonStart_pressed)

        self.row1CBoxA.addItems([str(r1a+1) for r1a in range(1000)])
        self.row2CBoxA.addItems([str(r2a+1) for r2a in range(1000)])
        self.column1CBoxA.addItems([get_column_letter(c1a+1) for c1a in range(1000)])
        self.column2CBoxA.addItems([get_column_letter(c2a+1) for c2a in range(1000)])

        self.row1CBoxB.addItems([str(r2b+1) for r2b in range(1000)])
        self.row2CBoxB.addItems([str(r2b+1) for r2b in range(1000)])
        self.column1CBoxB.addItems([get_column_letter(r2b+1) for r2b in range(1000)])
        self.column2CBoxB.addItems([get_column_letter(r2b+1) for r2b in range(1000)])

        # 환경설정
        self.settings = QSettings('config.ini', QSettings.IniFormat)

        # 테스트용
        try:
            print("try!")
            self.row1CBoxA.setCurrentText(self.settings.value('cell/row1CBoxA'))
            self.row2CBoxA.setCurrentText(self.settings.value('cell/row2CBoxA'))
            self.column1CBoxA.setCurrentText(self.settings.value('cell/column1CBoxA'))
            self.column2CBoxA.setCurrentText(self.settings.value('cell/column2CBoxA'))
            self.row1CBoxB.setCurrentText(self.settings.value('cell/row1CBoxB'))
            self.row2CBoxB.setCurrentText(self.settings.value('cell/row2CBoxB'))
            self.column1CBoxB.setCurrentText(self.settings.value('cell/column1CBoxB'))
            self.column2CBoxB.setCurrentText(self.settings.value('cell/column2CBoxB'))
        except:
            print("except!")
            self.row1CBoxA.setCurrentText("2")
            self.row2CBoxA.setCurrentText("13")
            self.column1CBoxA.setCurrentText("B")
            self.column2CBoxA.setCurrentText("I")
            self.row1CBoxB.setCurrentText("4")
            self.row2CBoxB.setCurrentText("19")
            self.column1CBoxB.setCurrentText("E")
            self.column2CBoxB.setCurrentText("O")

        self.setWindowTitle("엑셀값 연결")


        # 이전 실행 값 가져오기


        # self.row1CBoxA.currentIndexChanged.connect(self.selectionChanged)

        # 메뉴바 액션
        self.actionSite.triggered.connect(lambda: webbrowser.open('https://convertio.co/kr/xls-xlsx/'))


    def buttonA_pressed(self):
        self.file1, check = QFileDialog.getOpenFileName(None,
                                                       'Select file',
                                                       './',
                                                       'Excel Files (*.xlsx *.xls)')
        if check:
            try:
                # xls 파일일 경우
                if self.file1[-3:] == "xls":
                    # df = pd.read_excel(self.file1, header=None)
                    # df.to_excel(self.file1 + "x", index=False, header=False)
                    pe.save_book_as(file_name=self.file1, dest_file_name=self.file1+"x")
                    # 변수 동기화
                    self.file1 += 'x'
            except Exception as e:
                print('변환 실패')
                QMessageBox.about(self, "실패", "변환에 실패하였습니다.")

            file_name = self.file1.split('/')[-1]
            self.fileNameA.setText(file_name)
            self.sheetNameACBox.clear()
            self.sheetNameACBox.addItems(sheet_search(self.file1))

    def buttonB_pressed(self):
        self.file2, check = QFileDialog.getOpenFileName(None,
                                                  'Select file',
                                                  './',
                                                  'Excel Files (*.xlsx *.xls)')
        if check:
            try:
                if self.file2[-3:] == "xls":
                    # df = pd.read_excel(self.file2, header=None)
                    # df.to_excel(self.file2 + "x", index=False, header=False)
                    pe.save_book_as(file_name=self.file2, dest_file_name=self.file2 + "x")
                    self.file2 += 'x'
            except:
                print('변환 실패')
                QMessageBox.about(self, "실패", "변환에 실패하였습니다.")

            file_name = self.file2.split('/')[-1]
            self.fileNameB.setText(file_name)
            self.sheetNameBCBox.clear()
            self.sheetNameBCBox.addItems(sheet_search(self.file2))

    #     def selectionChanged(self):
    #     txt = self.cbo.currentText()
    #     idx = self.cbo.currentIndex()
    def buttonStart_pressed(self):
        try:
            # print("file1: ",self.file1)
            # print("file2: ",self.file2)
            print("실행 중..")
            buttonStart_pressed2(self.file1, self.file2, self.sheetNameACBox.currentText(),self.sheetNameBCBox.currentText(), self.row1CBoxA.currentText(), self.row2CBoxA.currentText(), self.column1CBoxA.currentText(), self.column2CBoxA.currentText(),
                                 self.row1CBoxB.currentText(), self.row2CBoxB.currentText(), self.column1CBoxB.currentText(), self.column2CBoxB.currentText())
            # self.settings.setValue('row1CBoxA', self.row1CBoxA.currentText())
            # self.settings.setValue('row2CBoxA', self.row2CBoxA.currentText())
            # self.settings.setValue('column1CBoxA', self.column1CBoxA.currentText())
            # self.settings.setValue('column2CBoxA', self.column2CBoxA.currentText())
            # self.settings.setValue('row1CBoxB', self.row1CBoxB.currentText())
            # self.settings.setValue('row2CBoxB', self.row2CBoxB.currentText())
            # self.settings.setValue('column1CBoxB', self.column1CBoxB.currentText())
            # self.settings.setValue('column2CBoxB', self.column2CBoxB.currentText())
            QMessageBox.about(self, "완료", "실행이 완료되었습니다.\n" + result_file_list[0])
        except:
            print("start error")
            # alert
            QErrorMessage.showMessage(self,'실행에 실패하였습니다.')

        return


    def closeEvent(self, event):
        print("Closed", event)
        self.settings.setValue('cell/row1CBoxA', self.row1CBoxA.currentText())
        self.settings.setValue('cell/row2CBoxA', self.row2CBoxA.currentText())
        self.settings.setValue('cell/column1CBoxA', self.column1CBoxA.currentText())
        self.settings.setValue('cell/column2CBoxA', self.column2CBoxA.currentText())
        self.settings.setValue('cell/row1CBoxB', self.row1CBoxB.currentText())
        self.settings.setValue('cell/row2CBoxB', self.row2CBoxB.currentText())
        self.settings.setValue('cell/column1CBoxB', self.column1CBoxB.currentText())
        self.settings.setValue('cell/column2CBoxB', self.column2CBoxB.currentText())

if __name__ == '__main__':
    app = QApplication(sys.argv)
    # app.setStyleSheet(style)
    ex = MyMainWindow()
    ex.show()
    sys.exit(app.exec())

