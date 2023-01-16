import openpyxl
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import time
import xlrd
result_file_list = []
def sheet_search(file):
    """ 모든 시트를 찾아줌 """
    wb = openpyxl.load_workbook(file)
    worksheets_list = []
    for sheet in wb.worksheets:
        worksheets_list.append(sheet.title)

    return worksheets_list


def buttonStart_pressed2(file1, file2, sheetA, sheetB, row1A, row2A, col1A, col2A, row1B, row2B, col1B, col2B):
    result_file_list.clear()

    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    sheet1 = wb1[sheetA]
    sheet2 = wb2[sheetB]

    data1 = sheet1[str(col1A)+str(row1A):str(col1A)+str(row2A)]
    data2 = sheet1[str(col2A)+str(row1A):str(col2A)+str(row2A)]

    data3 = sheet2[str(col1B)+str(row1B):str(col1B)+str(row2B)]
    data4 = sheet2[str(col2B)+str(row1B):str(col2B)+str(row2B)]

    for row in data1:
        for row2 in data3:
            # print(row2[4].value)
            # if row2 !="":
            if row[0].value is not None:
                # print(row[1].value)
                # print(row2[4].value)
                if row[0].value == row2[0].value:

                    print("!", row[0], " = ", row2[0], row2[0].value)
                    print(row[0].column_letter, row[0].row)
                    print(row2[0].column_letter, row2[0].row)

                    sheet2[str(col2B)+str(row2[0].row)].value =  sheet1[str(col2A) + str(row[0].row)].value



    result_file = file2.split('.xlsx')[0] + "(result).xlsx"
    wb2.save(result_file)
    print(" 실행 완료 ")
    result_file_list.append(result_file)


    # print("file name: ",file1, file2)
    # print("sheet name: ",sheetA, sheetB)
    # print("rowA: ",row1A,row2A)
    # print("colA: ",col1A,col2A)
    # print("rowB: ",row1B,row2B)
    # print("colB: ",col1B,col2B)
